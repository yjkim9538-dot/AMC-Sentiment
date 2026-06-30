# 운용사 배포용 엑셀 양식 생성 (디자인 정돈 + 예시 정리 + 입력 편의/보호)
# 실행: python3 build_template.py  → public/templates/AMC_컨센서스_양식.xlsx
#
# 설계:
#  - 각 시트: [1]제목 [2]안내 [3]헤더 [4]예시(옅은 노랑, 잠금·참고용) [5~]입력칸(잠금 해제)
#  - 방향성/의견 드롭다운, 목표값 천단위 표시
#  - 시트 보호: 입력칸만 편집 가능(암호 없음 → 필요 시 검토 탭에서 해제 가능)
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.comments import Comment

OUT = os.path.join(os.path.dirname(__file__), '..', 'public', 'templates', 'AMC_컨센서스_양식.xlsx')
os.makedirs(os.path.dirname(OUT), exist_ok=True)

NAVY = '1A2A4A'
HEADBLUE = '25406B'
AMBER = 'FFF7E6'
ZEBRA = 'F7F9FC'
GREY = '8A8A8A'

TITLE_FONT = Font(color='FFFFFF', bold=True, size=14)
NOTE_FONT = Font(color='5A6678', size=9)
HEAD_FONT = Font(color='FFFFFF', bold=True, size=11)
EX_FONT = Font(color=GREY, italic=True, size=10)
GUIDE_TITLE = Font(color=NAVY, bold=True, size=15)
GUIDE_FONT = Font(color='3A4452', size=11)

# 조건부 서식 색상 (국내 증시 관례: 강세·매수=빨강↑ / 약세·매도=파랑↓ / 중립=회색)
BULL_FILL = PatternFill('solid', fgColor='FCE4E4'); BULL_FONT = Font(color='C0392B', bold=True, size=10)
NEUT_FILL = PatternFill('solid', fgColor='ECEEF1'); NEUT_FONT = Font(color='5A6678', size=10)
BEAR_FILL = PatternFill('solid', fgColor='E4ECFB'); BEAR_FONT = Font(color='1F5FBF', bold=True, size=10)
WARN_FILL = PatternFill('solid', fgColor='FFD4D4')  # 목표 하단>상단 등 입력 오류 경고

THIN = Side(style='thin', color='D7DEE8')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
UNLOCK = Protection(locked=False)

wb = Workbook()


def build_sheet(ws, title, headers, widths, example, dropdowns, num_cols, input_rows,
                prefill=None, left_cols=(), band=None):
    ncol = len(headers)
    last = chr(ord('A') + ncol - 1)

    # [1] 제목
    ws.merge_cells(f'A1:{last}1')
    c = ws['A1']; c.value = title; c.fill = PatternFill('solid', fgColor=NAVY)
    c.font = TITLE_FONT; c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 30

    # [2] 안내
    ws.merge_cells(f'A2:{last}2')
    c = ws['A2']
    c.value = '방향성·의견은 드롭다운에서 선택하세요. 옅은 노란색 행은 작성 예시(참고용)이며 입력은 그 아래 칸에 하시면 됩니다.'
    c.font = NOTE_FONT; c.alignment = LEFT
    ws.row_dimensions[2].height = 18

    # [3] 헤더
    for i, h in enumerate(headers):
        cell = ws.cell(row=3, column=i + 1, value=h)
        cell.fill = PatternFill('solid', fgColor=HEADBLUE)
        cell.font = HEAD_FONT; cell.alignment = CENTER; cell.border = BORDER
    ws.row_dimensions[3].height = 30

    # [4] 예시 행 (잠금, 옅은 노랑)
    for i, v in enumerate(example):
        cell = ws.cell(row=4, column=i + 1, value=v)
        cell.fill = PatternFill('solid', fgColor=AMBER)
        cell.font = EX_FONT; cell.border = BORDER
        cell.alignment = LEFT if (i in left_cols) else CENTER
    ws.row_dimensions[4].height = 28
    ws['A4'].comment = Comment('예시 행입니다(참고용). 입력은 아래 칸에 하세요.', 'guide')

    # [5~] 입력 칸 (잠금 해제) + 선택 사전입력(prefill)
    start = 5
    for r in range(start, start + input_rows):
        ws.row_dimensions[r].height = 26
        for cidx in range(1, ncol + 1):
            cell = ws.cell(row=r, column=cidx)
            cell.border = BORDER
            cell.alignment = LEFT if (cidx - 1 in left_cols) else CENTER
            if (r - start) % 2 == 1:
                cell.fill = PatternFill('solid', fgColor=ZEBRA)
            # 숫자 칸 천단위 표시
            if (cidx - 1) in num_cols:
                cell.number_format = '#,##0'
            # 기본: 입력 가능하게 잠금 해제
            cell.protection = UNLOCK
        # 해외 시장명/기준지수 사전입력(잠금 유지)
        if prefill:
            pf = prefill.get(r - start)
            if pf:
                for cidx, val in pf.items():
                    cell = ws.cell(row=r, column=cidx + 1, value=val)
                    cell.protection = Protection(locked=True)
                    cell.font = Font(bold=True, color=NAVY, size=10)
                    cell.fill = PatternFill('solid', fgColor='EEF3FB')

    # 열 너비
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # 틀 고정(제목·안내·헤더)
    ws.freeze_panes = 'A4'
    ws.sheet_view.showGridLines = False

    # 드롭다운(입력 범위)
    for col_letter, options in dropdowns:
        dv = DataValidation(type='list', formula1='"%s"' % ','.join(options), allow_blank=True)
        dv.error = '목록에서 선택: ' + ' / '.join(options)
        dv.errorTitle = '입력 값 확인'
        dv.prompt = ' / '.join(options) + ' 중 선택'
        dv.promptTitle = '선택'
        ws.add_data_validation(dv)
        dv.add(f'{col_letter}{start}:{col_letter}{start + input_rows - 1}')

    end = start + input_rows - 1

    # 조건부 서식 1 — 방향성/의견 값에 따라 셀 색상(강세·매수=빨강 / 중립=회색 / 약세·매도=파랑)
    for col_letter, options in dropdowns:
        if len(options) != 3:
            continue
        rng = f'{col_letter}{start}:{col_letter}{end}'
        for opt, (fl, fo) in zip(options, [(BULL_FILL, BULL_FONT), (NEUT_FILL, NEUT_FONT), (BEAR_FILL, BEAR_FONT)]):
            ws.conditional_formatting.add(
                rng, CellIsRule(operator='equal', formula=['"%s"' % opt], fill=fl, font=fo))

    # 조건부 서식 2 — 목표 하단>상단 이면 두 칸을 빨갛게(입력 오류 경고)
    if band:
        lo, hi = band
        f = f'AND(${lo}{start}<>"",${hi}{start}<>"",${hi}{start}<${lo}{start})'
        ws.conditional_formatting.add(f'{lo}{start}:{hi}{end}', FormulaRule(formula=[f], fill=WARN_FILL))

    # 시트 보호 — 입력칸(잠금 해제)만 편집 가능. 암호 없음(검토 탭에서 해제 가능).
    ws.protection.sheet = True
    ws.protection.selectLockedCells = True
    ws.protection.selectUnlockedCells = True


# ── 시트 0: 작성안내 ──
guide = wb.active
guide.title = '작성안내'
guide.sheet_view.showGridLines = False
guide['A1'] = '증시 컨센서스 작성 안내'
guide['A1'].font = GUIDE_TITLE
lines = [
    '',
    '군인공제회 증권운용1팀 — 위탁운용사(AMC) 증시 컨센서스 제출 양식입니다.',
    '',
    '작성 방법',
    '  1. 아래 3개 시트를 채워 주세요:  [국내시장]  [국내종목]  [해외]',
    '  2. 각 시트의 옅은 노란색 행은 "작성 예시"입니다(참고용, 수정 불필요).',
    '     실제 입력은 그 아래 흰색 칸에 하시면 됩니다.',
    '  3. 방향성(강세/중립/약세), 의견(매수/중립/매도)은 칸을 클릭하면 나오는 드롭다운에서 선택합니다.',
    '     선택하면 칸 색이 바뀝니다 — 강세·매수=빨강, 중립=회색, 약세·매도=파랑.',
    '  4. 목표밴드는 지수의 예상 하단~상단을 숫자로 입력합니다(천단위 콤마 자동).',
    '     상단이 하단보다 작으면 두 칸이 빨갛게 표시되니 값을 확인하세요.',
    '  5. [국내종목]은 운용사가 자유 선정한 Top Pick을 적습니다.',
    '  6. [해외]는 미국/일본/베트남/인도/ACWI/선진국 6개 시장이 미리 채워져 있습니다.',
    '     각 행에 운용사명·작성일·방향성·목표·Pro/Con만 채우시면 됩니다.',
    '     6개 외 추가 시장은 그 아래 빈 칸에 직접 입력하면 됩니다.',
    '',
    '입력 보호',
    '  · 입력 칸 외에는 잠겨 있어 실수로 양식이 깨지지 않습니다.',
    '  · 행을 추가하는 등 양식을 바꿔야 하면: 상단 [검토] 탭 → [시트 보호 해제](암호 없음).',
    '',
    '제출',
    '  · 작성 후 파일을 그대로 담당자에게 회신하시면 됩니다(파일명 자유).',
    '  · 운용사명은 각 시트의 "첫 입력 행"에만 적으면 됩니다. 이어지는 행은 자동으로 같은',
    '    운용사로 인식됩니다. (한 파일은 한 운용사가 작성하는 것을 기준으로 합니다.)',
]
for i, t in enumerate(lines, start=2):
    cell = guide.cell(row=i, column=1, value=t)
    cell.font = GUIDE_FONT; cell.alignment = LEFT
guide.column_dimensions['A'].width = 100

# ── 국내시장 ──
m = wb.create_sheet('국내시장')
build_sheet(
    m, '국내 증시 전망 (운용사당 1행)',
    ['운용사명', '작성일', '방향성', 'KOSPI 목표 하단', 'KOSPI 목표 상단', 'Pro (긍정사유)', 'Con (부정사유)'],
    {'A': 22, 'B': 13, 'C': 11, 'D': 15, 'E': 15, 'F': 42, 'G': 42},
    ['(예시) 미래에셋자산운용', '2026-06-20', '강세', 2750, 3050,
     '반도체 업황 회복, 외국인 순매수 전환', '미국 금리 인하 지연, 중국 둔화'],
    dropdowns=[('C', ['강세', '중립', '약세'])],
    num_cols=[3, 4], input_rows=6, left_cols=(0, 5, 6), band=('D', 'E'),
)
m['C3'].comment = Comment('강세 / 중립 / 약세 중 선택', 'guide')
m['B3'].comment = Comment('날짜 형식 예: 2026-06-20', 'guide')
m['A3'].comment = Comment('운용사명은 이 행(첫 입력칸)에만 적으면 됩니다. 한 운용사가 한 파일을 작성합니다.', 'guide')

# ── 국내종목 ──
s = wb.create_sheet('국내종목')
build_sheet(
    s, '국내 종목 매수/매도 의견 (자유 선정 Top Pick)',
    ['운용사명', '종목명', '의견', '사유'],
    {'A': 22, 'B': 18, 'C': 10, 'D': 60},
    ['(예시) 미래에셋자산운용', '삼성전자', '매수', 'HBM 경쟁력 회복, 메모리 업황 반등'],
    dropdowns=[('C', ['매수', '중립', '매도'])],
    num_cols=[], input_rows=30, left_cols=(0, 3),
)
s['C3'].comment = Comment('매수 / 중립 / 매도 중 선택', 'guide')
s['A3'].comment = Comment('운용사명은 첫 종목 행에만 적으면 됩니다. 아래 종목들은 자동으로 같은 운용사로 인식됩니다.', 'guide')

# ── 해외 ── (6개 시장 사전입력)
o = wb.create_sheet('해외')
markets = [('미국', 'S&P500'), ('일본', 'Nikkei225'), ('베트남', 'VN-Index'),
           ('인도', 'Nifty50'), ('ACWI', 'MSCI ACWI'), ('선진국', 'MSCI World')]
prefill = {i: {2: mk, 3: idx} for i, (mk, idx) in enumerate(markets)}  # 입력칸 기준 0,1,2... 행에 시장/기준지수
build_sheet(
    o, '해외 증시 전망 (시장별)',
    ['운용사명', '작성일', '시장', '기준지수', '방향성', '목표 하단', '목표 상단', 'Pro (긍정사유)', 'Con (부정사유)'],
    {'A': 22, 'B': 13, 'C': 10, 'D': 13, 'E': 11, 'F': 12, 'G': 12, 'H': 38, 'I': 38},
    ['(예시) 미래에셋자산운용', '2026-06-20', '미국', 'S&P500', '강세', 5400, 6000,
     'AI 투자 사이클 지속', '밸류에이션 부담'],
    dropdowns=[('E', ['강세', '중립', '약세'])],
    num_cols=[5, 6], input_rows=10, prefill=prefill, left_cols=(0, 7, 8), band=('F', 'G'),
)
o['E3'].comment = Comment('강세 / 중립 / 약세 중 선택', 'guide')
o['B3'].comment = Comment('날짜 형식 예: 2026-06-20', 'guide')
o['A3'].comment = Comment('운용사명은 첫 행에만 적으면 됩니다(아래 행 자동 인식).', 'guide')
o['C11'].comment = Comment('기본 6개 시장 외에 추가할 시장이 있으면 이 아래 빈 칸에 직접 입력하세요.', 'guide')

wb.save(OUT)
print('saved:', os.path.abspath(OUT))
